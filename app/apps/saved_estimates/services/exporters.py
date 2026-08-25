import csv
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font


def estimate_csv_response(estimate):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="{estimate.reference}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(['Reference', estimate.reference])
    writer.writerow(['Client', estimate.client.code])
    writer.writerow(['Created by', estimate.created_by_label])
    writer.writerow(['Created at', estimate.created_at.isoformat()])
    writer.writerow(['Destination', estimate.destination_label])
    writer.writerow([])
    writer.writerow(['Carrier', 'Service', 'Estimate ex GST', 'Status'])
    for result in estimate.result_snapshot:
        writer.writerow([
            result.get('carrier', ''),
            result.get('service', ''),
            result.get('estimate_ex_gst', ''),
            result.get('status', ''),
        ])
    return response


def estimate_xlsx_response(estimate):
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'
    summary.append(['Field', 'Value'])
    summary['A1'].font = Font(bold=True)
    summary['B1'].font = Font(bold=True)
    for field, value in [
        ('Reference', estimate.reference),
        ('Client', estimate.client.code),
        ('Created by', estimate.created_by_label),
        ('Created at', estimate.created_at.isoformat()),
        ('Destination', estimate.destination_label),
        ('Total weight (kg)', estimate.total_weight_kg),
        ('Total cubic (m3)', estimate.total_cubic_m3),
        ('Best estimate ex GST', estimate.best_estimate_ex_gst),
    ]:
        summary.append([field, value])
    summary.column_dimensions['A'].width = 24
    summary.column_dimensions['B'].width = 42

    items = workbook.create_sheet('Items')
    item_headers = [
        'SKU', 'Quantity', 'Type', 'Length (m)', 'Width (m)', 'Height (m)',
        'Weight (kg)', 'Cubic (m3)',
    ]
    items.append(item_headers)
    for cell in items[1]:
        cell.font = Font(bold=True)
    for line in estimate.input_snapshot.get('lines', []):
        items.append([
            line.get('sku', ''),
            line.get('quantity', ''),
            line.get('freight_type', ''),
            line.get('length_m', ''),
            line.get('width_m', ''),
            line.get('height_m', ''),
            line.get('weight_kg', ''),
            line.get('cubic_m3', ''),
        ])

    options = workbook.create_sheet('Freight options')
    option_headers = [
        'Carrier', 'Service', 'Estimate ex GST', 'Status', 'Zone',
        'Chargeable weight', 'Fuel', 'Tailgate fee', 'Handling',
    ]
    options.append(option_headers)
    for cell in options[1]:
        cell.font = Font(bold=True)
    for result in estimate.result_snapshot:
        details = result.get('details') or {}
        options.append([
            result.get('carrier', ''),
            result.get('service', ''),
            result.get('estimate_ex_gst', ''),
            result.get('status', ''),
            details.get('zone', ''),
            details.get('chargeable_weight', ''),
            details.get('fuel', ''),
            details.get('tailgate_fee', ''),
            details.get('handling', ''),
        ])

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{estimate.reference}.xlsx"'
    )
    return response

