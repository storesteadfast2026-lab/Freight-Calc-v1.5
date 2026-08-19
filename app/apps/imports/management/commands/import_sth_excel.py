from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from apps.clients.models import Client, FreightCalculator
from apps.locations.models import Suburb, FromAddress
from apps.products.models import Product
from apps.carriers.models import Carrier, CarrierService, ClientCarrierConfig
from apps.rates.models import FreightZone, FreightRate, CarrierTailgateCharge
from apps.imports.models import ExternalDataFile
from apps.imports.services.fuel import calculate_sha256, reapply_active_fuel_rates


def d(value, default='0'):
    if value in (None, ''):
        return Decimal(default)
    try:
        text = str(value).replace('%', '').strip()
        val = Decimal(text)
        return val / Decimal('100') if '%' in str(value) else val
    except (InvalidOperation, ValueError):
        return Decimal(default)


def s(value):
    return str(value or '').strip()


def yes(value) -> bool:
    return s(value).upper() == 'YES'


class Command(BaseCommand):
    help = 'Import main STH workbook sheets into PostgreSQL tables using cached Excel values.'

    def add_arguments(self, parser):
        parser.add_argument('workbook_path')
        parser.add_argument('--client', default='STH')
        parser.add_argument('--replace', action='store_true', help='Replace imported data for the client before importing.')
        parser.add_argument(
            '--fuel-source',
            choices=['active', 'workbook'],
            default='active',
            help=(
                'active: reapply the latest ACTIVE Admin fuel file after workbook import; '
                'workbook: keep the fuel levy cached in the workbook (historical validation only).'
            ),
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options['workbook_path'])
        if not path.exists():
            raise CommandError(f'Workbook not found: {path}')

        client, _ = Client.objects.get_or_create(code=options['client'], defaults={'name': 'Stenhoj Australia'})
        FreightCalculator.objects.get_or_create(client=client, name='STH Freight Calculator', version='V2026.R2')
        FromAddress.objects.get_or_create(client=client, name='Default STH FROM', defaults={'suburb': '', 'state': '', 'postcode': '', 'is_default': True})

        if options['replace']:
            self.replace_client_data(client)

        # data_only=True reads cached Excel values. Fuel levy is a legacy bootstrap only;
        # an active Admin fuel dataset is reapplied after ClientCarrierConfig is rebuilt.
        wb = load_workbook(path, data_only=True, read_only=True)
        summary = {}
        summary['suburbs'] = self.import_suburbs(wb, client)
        summary['products'] = self.import_products(wb, client)
        summary['carriers'] = self.import_carriers(wb, client)
        if options['fuel_source'] == 'active':
            summary['fuel'] = reapply_active_fuel_rates(client)
        else:
            summary['fuel'] = {'source': 'LEGACY_WORKBOOK', 'configs_reapplied': 0}
        summary['zones'] = self.import_zones(wb, client)
        summary['rates'] = self.import_rates(wb, client)
        summary['tailgate'] = self.import_tailgate(wb, client)

        ExternalDataFile.objects.create(
            client=client,
            file_type='WORKBOOK',
            original_filename=path.name,
            stored_path=str(path),
            source_method='COMMAND',
            file_size_bytes=path.stat().st_size,
            sha256=calculate_sha256(path.read_bytes()),
            status='IMPORTED',
            last_imported_at=timezone.now(),
            import_summary=summary,
        )
        self.stdout.write(self.style.SUCCESS(f'Imported workbook for {client.code}: {summary}'))

    def replace_client_data(self, client):
        """Clear imported tables for the selected client."""
        FreightRate.objects.filter(client=client).delete()
        FreightZone.objects.filter(client=client).delete()
        CarrierTailgateCharge.objects.filter(client=client).delete()
        ClientCarrierConfig.objects.filter(client=client).delete()
        Product.objects.filter(client=client).delete()
        # Preserve FUEL history and the active web/admin fuel dataset.
        ExternalDataFile.objects.filter(client=client).exclude(file_type='FUEL').delete()
        # Suburbs are global Australian reference data; keep existing and update/create.

    def import_suburbs(self, wb, client):
        if 'SUBURBS' not in wb.sheetnames:
            return 0
        ws = wb['SUBURBS']
        count = 0
        batch = []
        existing = set(Suburb.objects.values_list('normalized_key', 'postcode'))
        for row in ws.iter_rows(min_row=3, values_only=True):
            state = s(row[3] if len(row) > 3 else '')
            suburb = s(row[4] if len(row) > 4 else '')
            postcode = s(row[5] if len(row) > 5 else '')
            if not state or not suburb or not postcode:
                continue
            key = f'{state}{suburb}'.upper().strip()
            if (key, postcode) in existing:
                continue
            batch.append(Suburb(suburb_name=suburb, state=state, postcode=postcode, normalized_key=key))
            existing.add((key, postcode))
            count += 1
            if len(batch) >= 2000:
                Suburb.objects.bulk_create(batch, ignore_conflicts=True)
                batch.clear()
        if batch:
            Suburb.objects.bulk_create(batch, ignore_conflicts=True)
        return count

    def import_products(self, wb, client):
        if 'SKUs' not in wb.sheetnames:
            return 0
        ws = wb['SKUs']
        count = 0
        products = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            sku = s(row[0] if len(row) > 0 else '')
            if not sku or sku.upper() == 'SKU':
                continue
            products.append(Product(
                client=client,
                sku=sku,
                name=s(row[1] if len(row) > 1 else '') or sku,
                description=s(row[2] if len(row) > 2 else ''),
                length_m=d(row[4] if len(row) > 4 else 0),
                width_m=d(row[5] if len(row) > 5 else 0),
                height_m=d(row[6] if len(row) > 6 else 0),
                weight_kg=d(row[7] if len(row) > 7 else 0),
                cubic_m3=d(row[8] if len(row) > 8 else 0),
                freight_type=(s(row[9] if len(row) > 9 else 'P') or 'P')[:1],
                source_row=idx,
                active=True,
            ))
            count += 1
        Product.objects.bulk_create(products, ignore_conflicts=True, batch_size=1000)
        return count

    def _carrier_service(self, carrier_code, service_code):
        carrier, _ = Carrier.objects.get_or_create(code=carrier_code, defaults={'name': carrier_code})
        service, _ = CarrierService.objects.get_or_create(carrier=carrier, service_code=service_code, defaults={'service_name': service_code})
        return carrier, service

    def import_carriers(self, wb, client):
        if 'FuelSurcharge' not in wb.sheetnames:
            return 0
        ws = wb['FuelSurcharge']
        count = 0
        handling_charge = d(wb['SettingFlags']['E20'].value if 'SettingFlags' in wb.sheetnames else 0)
        for row in ws.iter_rows(min_row=6, max_row=135, values_only=True):
            # FuelSurcharge columns: G Order, H Carrier, I Ratecard, J Service, K Fuel Levy, ... AD PostC Zones.
            carrier_code = s(row[7] if len(row) > 7 else '')
            service_code = s(row[9] if len(row) > 9 else '')
            if not carrier_code or not service_code:
                continue
            carrier, service = self._carrier_service(carrier_code, service_code)
            base_status = s(row[25] if len(row) > 25 else 'L') or 'L'
            order_ready_formula_value = s(row[16] if len(row) > 16 else 'YES')
            order_ready_rule = 'WOODVILLE_NORTH_ONLY' if carrier_code == 'CUST' else 'GOOD'
            ClientCarrierConfig.objects.update_or_create(
                client=client,
                carrier_service=service,
                defaults={
                    'customer_code': 'STH',
                    'ratecard': s(row[8] if len(row) > 8 else ''),
                    # Legacy bootstrap only. If an active FUEL file exists, it is reapplied
                    # immediately after carrier configs are rebuilt.
                    'fuel_levy': d(row[10] if len(row) > 10 else 0),
                    'fuel_levy_source': 'LEGACY_WORKBOOK',
                    'fuel_levy_updated_at': timezone.now(),
                    'fuel_data_file': None,
                    'extra_surcharge': d(row[11] if len(row) > 11 else 0),
                    'uprate': d(row[12] if len(row) > 12 else 0),
                    'tailgate_enabled': yes(row[13] if len(row) > 13 else ''),
                    'warehouse_handling_enabled': yes(row[14] if len(row) > 14 else ''),
                    'fixed_handling_charge': handling_charge,
                    'cubic_conversion': d(row[15] if len(row) > 15 else 0),
                    'pallet_enabled': yes(row[17] if len(row) > 17 else ''),
                    'carton_enabled': yes(row[18] if len(row) > 18 else ''),
                    'hand_unload_enabled': yes(row[19] if len(row) > 19 else ''),
                    'subzone_enabled': yes(row[20] if len(row) > 20 else ''),
                    'area_enabled': yes(row[21] if len(row) > 21 else ''),
                    'overlength_enabled': yes(row[22] if len(row) > 22 else ''),
                    'zone_enabled': yes(row[27] if len(row) > 27 else ''),
                    'empty_rate_enabled': yes(row[28] if len(row) > 28 else ''),
                    'postcode_zones_enabled': yes(row[29] if len(row) > 29 else ''),
                    'base_status': base_status,
                    'order_ready_rule': order_ready_rule,
                    'active': base_status == 'L',
                }
            )
            count += 1
        return count

    def import_zones(self, wb, client):
        if 'ZONES' not in wb.sheetnames:
            return 0
        ws = wb['ZONES']
        count = 0
        batch = []
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # ZONES columns: D Carrier, E Service, F Suburb, G State, H Postcode, I Zone, J Subzone, K Area.
            carrier_code = s(row[3] if len(row) > 3 else '')
            service_code = s(row[4] if len(row) > 4 else '')
            suburb = s(row[5] if len(row) > 5 else '')
            state = s(row[6] if len(row) > 6 else '')
            postcode = s(row[7] if len(row) > 7 else '')
            zone = s(row[8] if len(row) > 8 else '')
            subzone = s(row[9] if len(row) > 9 else '')
            area = s(row[10] if len(row) > 10 else '')
            if not carrier_code or not service_code or not suburb or not state:
                continue
            _, service = self._carrier_service(carrier_code, service_code)
            zone_obj = FreightZone(client=client, carrier_service=service, suburb=suburb, state=state, postcode=postcode, zone=zone, subzone=subzone, area=area, source_row=idx)
            zone_obj.lookup_key_suburb = f'{service.excel_key}{suburb}{state}'.upper().strip()
            zone_obj.lookup_key_postcode = f'{service.excel_key}{postcode}'.upper().strip()
            batch.append(zone_obj)
            count += 1
            if len(batch) >= 2000:
                FreightZone.objects.bulk_create(batch, batch_size=2000)
                batch.clear()
        if batch:
            FreightZone.objects.bulk_create(batch, batch_size=2000)
        return count

    def import_rates(self, wb, client):
        if 'RATES' not in wb.sheetnames:
            return 0
        ws = wb['RATES']
        count = 0
        batch = []
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            carrier_code = s(row[2] if len(row) > 2 else '')
            service_code = s(row[3] if len(row) > 3 else '')
            if not carrier_code or not service_code:
                continue
            _, service = self._carrier_service(carrier_code, service_code)
            rate_obj = FreightRate(
                client=client, carrier_service=service,
                zone=s(row[4] if len(row)>4 else ''), subzone=s(row[5] if len(row)>5 else ''), area=s(row[6] if len(row)>6 else ''),
                weight_break=s(row[7] if len(row)>7 else ''), freight_type=s(row[8] if len(row)>8 else ''),
                minimum_charge=d(row[9] if len(row)>9 else 0), basic_charge=d(row[10] if len(row)>10 else 0),
                per_subsequent_basic=d(row[11] if len(row)>11 else 0), per_kg=d(row[12] if len(row)>12 else 0),
                overweight_charge=d(row[13] if len(row)>13 else 0), remote_charge=d(row[14] if len(row)>14 else 0),
                offshore_charge=d(row[15] if len(row)>15 else 0), overlength_charge=d(row[16] if len(row)>16 else 0),
                customer_code=s(row[17] if len(row)>17 else 'STH') or 'STH', margin=d(row[18] if len(row)>18 else 0),
                source_row=idx
            )
            rate_obj.lookup_key = ''.join([
                service.excel_key,
                rate_obj.zone or '', rate_obj.subzone or '', rate_obj.area or '',
                rate_obj.weight_break or '', rate_obj.customer_code or '', rate_obj.freight_type or ''
            ]).strip().upper()
            batch.append(rate_obj)
            count += 1
            if len(batch) >= 2000:
                FreightRate.objects.bulk_create(batch, batch_size=2000)
                batch.clear()
        if batch:
            FreightRate.objects.bulk_create(batch, batch_size=2000)
        return count

    def import_tailgate(self, wb, client):
        if 'SettingFlags' not in wb.sheetnames:
            return 0
        ws = wb['SettingFlags']
        count = 0
        for row in ws.iter_rows(min_row=34, max_row=52, values_only=True):
            carrier_code = s(row[2] if len(row) > 2 else '')
            if not carrier_code:
                continue
            carrier, _ = Carrier.objects.get_or_create(code=carrier_code, defaults={'name': carrier_code})
            CarrierTailgateCharge.objects.update_or_create(
                client=client, carrier=carrier,
                defaults={
                    'minimum_charge': d(row[4] if len(row)>4 else 0),
                    'per_subsequent_charge': d(row[5] if len(row)>5 else 0),
                    'hand_unload_charge': d(row[6] if len(row)>6 else 0),
                }
            )
            count += 1
        return count
