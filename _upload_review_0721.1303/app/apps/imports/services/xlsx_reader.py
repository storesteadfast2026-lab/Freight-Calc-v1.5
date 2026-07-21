from __future__ import annotations

import hashlib
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from apps.imports.models import ExternalDataFile


class SourceImportError(Exception):
    """Raised when a product/stock reference source cannot be read or validated."""


def read_file_bytes(external_file: ExternalDataFile) -> bytes:
    if external_file.uploaded_file:
        external_file.uploaded_file.open('rb')
        try:
            return external_file.uploaded_file.read()
        finally:
            external_file.uploaded_file.close()
    if external_file.stored_path:
        path = Path(external_file.stored_path)
        if path.exists() and path.is_file():
            return path.read_bytes()
    raise SourceImportError('The stored source file is not available.')


def calculate_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def normalize_header(value: Any) -> str:
    text = str(value or '').strip().lower()
    return re.sub(r'[^a-z0-9]+', '', text)


def value_to_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'True' if value else 'False'
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def normalize_sku(value: Any) -> str:
    text = value_to_text(value).strip()
    if not text or text == '0':
        return ''
    try:
        number = Decimal(text)
        if number == number.to_integral_value():
            return str(number.quantize(Decimal('1')))
    except InvalidOperation:
        pass
    return text.upper()


def json_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def parse_decimal(
    value: Any,
    *,
    field_label: str,
    row_number: int,
    errors: list[str],
    allow_blank: bool = True,
    non_negative: bool = True,
) -> Decimal | None:
    if value is None or str(value).strip() == '':
        if not allow_blank:
            errors.append(f'Row {row_number}: {field_label} is required.')
        return None
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError):
        errors.append(f'Row {row_number}: {field_label} is not numeric ({value!r}).')
        return None
    if non_negative and decimal_value < 0:
        errors.append(f'Row {row_number}: {field_label} cannot be negative ({value!r}).')
    return decimal_value


def _aliases_by_normalized_name(aliases: dict[str, tuple[str, ...]]) -> dict[str, str]:
    output: dict[str, str] = {}
    for logical_name, values in aliases.items():
        for value in values:
            output[normalize_header(value)] = logical_name
    return output


def _header_mapping(values: tuple[Any, ...], aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
    lookup = _aliases_by_normalized_name(aliases)
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        logical_name = lookup.get(normalize_header(value))
        if logical_name and logical_name not in mapping:
            mapping[logical_name] = index
    return mapping


def read_xlsx_records(
    content: bytes,
    *,
    preferred_sheet_names: tuple[str, ...],
    aliases: dict[str, tuple[str, ...]],
    required_fields: tuple[str, ...],
    max_header_scan_rows: int = 25,
) -> tuple[str, int, list[str], list[dict[str, Any]]]:
    if not content:
        raise SourceImportError('The uploaded Excel file is empty.')

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise SourceImportError(f'The uploaded file is not a readable .xlsx workbook: {exc}') from exc

    try:
        preferred = {normalize_header(name) for name in preferred_sheet_names}
        sheets = sorted(
            workbook.worksheets,
            key=lambda sheet: 0 if normalize_header(sheet.title) in preferred else 1,
        )

        selected = None
        best_match_count = -1
        for sheet in sheets:
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=max_header_scan_rows, values_only=True),
                start=1,
            ):
                mapping = _header_mapping(row, aliases)
                match_count = len(mapping)
                if all(field in mapping for field in required_fields):
                    selected = (sheet, row_number, row, mapping)
                    break
                if match_count > best_match_count:
                    best_match_count = match_count
            if selected:
                break

        if selected is None:
            expected = ', '.join(required_fields)
            raise SourceImportError(
                'Could not find a worksheet/header row with the required columns. '
                f'Expected logical fields: {expected}.'
            )

        sheet, header_row_number, header_values, mapping = selected
        headers = [value_to_text(value) or f'column_{index + 1}' for index, value in enumerate(header_values)]
        records: list[dict[str, Any]] = []

        for source_row_number, values in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            if not any(value not in (None, '') for value in values):
                continue

            record = {
                logical_name: values[column_index] if column_index < len(values) else None
                for logical_name, column_index in mapping.items()
            }
            raw_data = {
                headers[index]: json_safe_value(value)
                for index, value in enumerate(values[:len(headers)])
                if value not in (None, '')
            }
            record['_row_number'] = source_row_number
            record['_raw_data'] = raw_data
            records.append(record)

        if not records:
            raise SourceImportError('The workbook contains headers but no data rows.')
        return sheet.title, header_row_number, headers, records
    finally:
        workbook.close()
