from pathlib import Path
from openpyxl import load_workbook

workbook_path = Path("generated_excel_baselines/latest/STH_LIVE_BASELINE_20260706_130237.xlsx")

if not workbook_path.exists():
    raise SystemExit(f"Workbook not found: {workbook_path}")

wb = load_workbook(workbook_path, data_only=True, read_only=True)

if "Calculator" not in wb.sheetnames:
    raise SystemExit(f"Calculator sheet not found. Sheets: {wb.sheetnames}")

ws = wb["Calculator"]

carrier_codes = {"KTI", "TEAMEX", "TFMX", "STEA", "COCHRN", "MIPEC", "TEAMTAS", "CUST"}

print("Scanning Calculator sheet for likely output cells...")
print("=" * 80)

for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=30):
    values = []
    has_relevant = False

    for cell in row:
        value = cell.value

        if value is None:
            continue

        text = str(value).strip()

        if not text:
            continue

        if text in carrier_codes:
            has_relevant = True

        if any(code in text for code in carrier_codes):
            has_relevant = True

        if isinstance(value, (int, float)) and 10 <= value <= 10000:
            has_relevant = True

        values.append(f"{cell.coordinate}={text}")

    if has_relevant and values:
        print(" | ".join(values))
